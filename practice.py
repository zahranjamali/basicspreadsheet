import openpyxl

inv_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inv_file["Sheet1"]

product_list_dictionary = {}
total_inventory_value = {}
product_under_10_inv = {}

for product_row in range(2, product_list.max_row + 1):
    supply_names = product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    product_num = product_list.cell(product_row, 1).value
    inventory_value = product_list.cell(product_row, 5)
    if supply_names in product_list_dictionary:
        current_no_product = product_list_dictionary[supply_names]
        product_list_dictionary[supply_names] = current_no_product + 1
    else:
        product_list_dictionary[supply_names] = 1

    if supply_names in total_inventory_value:
        current_price = total_inventory_value[supply_names]
        total_inventory_value[supply_names] = current_price + inventory * price
    else:
        total_inventory_value[supply_names] = inventory * price

    if inventory < 10:
        product_under_10_inv[int(product_num)] = int(inventory)

    inventory_value.value = inventory * price

    inventory_value.value = inventory * price

print(f"total product list is {product_list_dictionary}")
print(f"total inventory value is {total_inventory_value}")
print(f"product under 10 are {product_under_10_inv}")

inv_file.save("inventory_with_total_value.xlsx")